

import pandas as pd
import numpy as np
import openpyxl
import os
from openpyxl import Workbook
from openpyxl import load_workbook

program_cikti_s=11
ders_cikti_s=6


program_cikti = list(range(1, (program_cikti_s)))  
ders_cikti = list(range(1, (ders_cikti_s)))  

iliskiler = [
    [1, 1, 0, 1, 1],
    [0, 0, 1, 0.2, 0],
    [0, 0, 0.5 , 1, 0],
    [0, 0, 0, 0.8, 1],
    [0, 1, 0, 1, 0],
    [1, 0, 0, 0, 1],
    [0, 0, 0, 1, 1],
    [0, 1, 1, 1, 1],
    [1, 0, 1, 1, 0],
    [0, 1, 1, 1, 0],
]




df = pd.DataFrame(iliskiler, columns=ders_cikti, index=program_cikti)


iliskiler_deg = []

katsayi = 1 / (ders_cikti_s - 1)

for index, row in df.iterrows():
    toplam = 0
    for value in row:  
        toplam += value * katsayi
    iliskiler_deg.append(round(toplam, 2))  


df["İlişki Değ."] = iliskiler_deg

file_name = "ders_program_cikti_iliskisi.xlsx"
if os.path.exists(file_name):
    print(f"{file_name} dosyası zaten mevcut. Yeni dosya oluşturulmadı.")
else:
    df.to_excel(file_name, sheet_name="Tablo 1", index_label="Prg Çıktı")
    print(f"{file_name} başarıyla oluşturuldu.")







ders_cikti_s = 5  
degerlendirme_sayisi_s = 6  


degerlendirme_sutunlari = ["Ödev1", "Ödev2", "Quiz", "Quiz4", "Vize", "Final"]
agirliklar = [10, 10, 10, 10, 20, 40]  

iliskiler_ders = [
     [1, 1, 0, 1, 1, 1],
    [0, 1, 1, 1, 0, 1],
    [1, 0, 1, 0, 1, 0],
    [0, 1, 0, 1, 1, 1],
    [1, 0, 1, 0, 0, 1]
]


df_degerlendirme = pd.DataFrame(iliskiler_ders, columns=degerlendirme_sutunlari, index=range(1, ders_cikti_s + 1))


agirliklar_dict = {degerlendirme_sutunlari[i]: agirliklar[i] for i in range(len(agirliklar))}
df_degerlendirme.loc["Ağırlıklar"] = agirliklar_dict


df_degerlendirme_without_weights = df_degerlendirme.drop(index="Ağırlıklar")  


df_degerlendirme_without_weights["Toplam"] = df_degerlendirme_without_weights.sum(axis=1)


df_degerlendirme["Toplam"] = df_degerlendirme_without_weights["Toplam"]


file_name = "ders_ciktisi_tablosu.xlsx"
if os.path.exists(file_name):
    print(f"{file_name} dosyası zaten mevcut. Yeni dosya oluşturulmadı.")
else:
   
    with pd.ExcelWriter(file_name) as writer:
        df_degerlendirme.to_excel(writer, sheet_name="ders_ciktisi_tablosu", index_label="Ders Çıktısı")

    print(f"{file_name} başarıyla oluşturuldu.")


file_name_ders_ciktisi = "ders_ciktisi_tablosu.xlsx"


if os.path.exists(file_name_ders_ciktisi):
    df_degerlendirme = pd.read_excel(file_name_ders_ciktisi, sheet_name="ders_ciktisi_tablosu", index_col="Ders Çıktısı")
    print(f"{file_name_ders_ciktisi} dosyasından veri okundu.")
else:
    print(f"{file_name_ders_ciktisi} bulunamadı. Lütfen dosyayı oluşturun.")
    exit()


agirliklar = df_degerlendirme.loc["Ağırlıklar"].values
degerlendirme_sutunlari = df_degerlendirme.columns[:-1] 


df_agirlikli_degerlendirme = pd.DataFrame(index=df_degerlendirme.index[:-1], columns=degerlendirme_sutunlari)  


for col in degerlendirme_sutunlari:
    for row in df_degerlendirme.index[:-1]: 
        
        new_value = (df_degerlendirme.loc[row, col] * agirliklar[degerlendirme_sutunlari.tolist().index(col)]) / 100
        df_agirlikli_degerlendirme.loc[row, col] = round(new_value, 2)  


df_agirlikli_degerlendirme["Toplam"] = df_agirlikli_degerlendirme.sum(axis=1)


file_name_agirlikli = "agirlikli_degerlendirme.xlsx"


if os.path.exists(file_name_agirlikli):
    print(f"{file_name_agirlikli} dosyası zaten mevcut. Yeni dosya oluşturulmadı.")
else:

    with pd.ExcelWriter(file_name_agirlikli) as writer:
        df_agirlikli_degerlendirme.to_excel(writer, sheet_name="Ağırlıklı Değerlendirme", index_label="Ders Çıktısı")
    print(f"{file_name_agirlikli} başarıyla oluşturuldu.")







file_name_dersanotlar = "DersANotlar.xlsx"
file_name_agirlikli = "agirlikli_degerlendirme.xlsx"
new_file_name = "agirlikli_ogrenci_notlari.xlsx" 


if os.path.exists(file_name_dersanotlar):
    df_dersanotlar = pd.read_excel(file_name_dersanotlar, sheet_name="Sayfa1", index_col="Ogrenci_No")
    print(f"{file_name_dersanotlar} dosyasından veri okundu.")
else:
    print(f"{file_name_dersanotlar} bulunamadı. Lütfen dosyayı oluşturun.") 

if os.path.exists(file_name_agirlikli):
    df_agirlikli = pd.read_excel(file_name_agirlikli, sheet_name="Ağırlıklı Değerlendirme", index_col="Ders Çıktısı")
    print(f"{file_name_agirlikli} dosyasından veri okundu.")
else:
    print(f"{file_name_agirlikli} bulunamadı. Lütfen dosyayı oluşturun.")
    exit()


with pd.ExcelWriter(new_file_name, engine='openpyxl') as writer: 
    
    start_row = 1

    for student in df_dersanotlar.index:
        student_data = df_dersanotlar.loc[student]

        student_table = pd.DataFrame(columns=df_agirlikli.columns, index=df_agirlikli.index)

        for column in df_agirlikli.columns:
            if column != "Toplam":  
                for row in df_agirlikli.index:
                    student_table.loc[row, column] = round(student_data[column] * df_agirlikli.loc[row, column], 2)

        
        student_table["Toplam"] = student_table.sum(axis=1)

        student_table["Max"] = student_table.apply(lambda row: sum([100 * df_agirlikli.loc[row.name, col] for col in df_agirlikli.columns[:-1]]), axis=1)

        student_table["Başarı Oranı"] = (student_table["Toplam"] / student_table["Max"]) * 100

        
        student_table.to_excel(writer, sheet_name="Tüm Öğrenciler", startcol=0, startrow=start_row, index_label=f"Öğrenci {student}")
        
        
        start_row += len(student_table) + 3  

        
        
    print(f"Tüm öğrencilerin tabloları tek bir sayfada başarıyla oluşturuldu.")



file_name_agirlikli_notlar = "agirlikli_ogrenci_notlari.xlsx"
file_name_ders_program_ciktisi = "ders_program_cikti_iliskisi.xlsx"
new_file_name_2 = "agirlikli_ders_program_ciktisi_notlari.xlsx"

if os.path.exists(file_name_ders_program_ciktisi):
    df_dersanotlar = pd.read_excel(file_name_ders_program_ciktisi, sheet_name="Tablo 1", index_col="Prg Çıktı")
    print(f"{file_name_ders_program_ciktisi} dosyasından veri okundu.")

     
    df_dersanotlar_sliced = df_dersanotlar.iloc[:, 0:-1]  

        
    column_lists = [df_dersanotlar_sliced[col].tolist() for col in df_dersanotlar_sliced.columns]

    print(column_lists)  
else:
    print(f"{file_name_ders_program_ciktisi} bulunamadı. Lütfen dosyayı oluşturun.")

if os.path.exists(file_name_agirlikli_notlar):
    
    
    df = pd.read_excel(file_name_agirlikli_notlar, sheet_name="Tüm Öğrenciler") 
    
   

    
    student_tables = []  
    student_numbers = []  
    
    basarim_orani_listesi = []  

   
    start_row = 0#
    while start_row < len(df):
        
        
        if isinstance(df.iloc[start_row, 0], str) and df.iloc[start_row, 0].startswith("Öğrenci"):
            
            student_number = df.iloc[start_row, 0]  
            student_numbers.append(student_number)  

            
            student_table = df.iloc[start_row + 1:start_row + 6]  
            student_tables.append(student_table) 

           
            basarim_orani = student_table.iloc[:, 9].tolist()  
            basarim_orani_listesi.append(basarim_orani)

           
            start_row += len(student_table) + 3  
        else:
            start_row += 1  

    
    """print(basarim_orani_listesi)"""
    result_lists=[]
    for first_item in basarim_orani_listesi:

        result=[]
        for second_item in column_lists:
            
            multiplied_values = [first_value * second_value for first_value, second_value in zip(first_item, second_item)]
            result.append(multiplied_values)
        result_lists.append(result)


for res in result_lists:
    print(res)



   
    with pd.ExcelWriter(new_file_name_2, engine='openpyxl') as writer:
        
        
        df = pd.read_excel("ders_program_cikti_iliskisi.xlsx", sheet_name="Tablo 1")
        
        start_row = 0
        program_cikti_sayilari = df['Prg Çıktı'].tolist() 
    
        print(program_cikti_sayilari)
        

            
        for student in student_numbers:
            
            sheet_name = "Ders Çıktısı"

            

            
            student_table = pd.DataFrame({
                'Prg Çıktı': program_cikti_sayilari,
                    
            })
            

           
            student_table.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 2, index=False)

            

            
        

            start_row += len(student_table) + 3
            

else:
    print(f"{file_name_agirlikli_notlar} bulunamadı. Lütfen dosyayı oluşturun.") 


    
workbook = load_workbook(new_file_name_2)
worksheet = workbook[sheet_name]


row = 2

for idx, student in enumerate(student_numbers):
    worksheet.cell(row=row, column=1, value=student)  
    worksheet.cell(row=row, column=2, value="Ders Çıktısı")  
    row += len(program_cikti) + 3  

workbook.save(new_file_name_2)
        